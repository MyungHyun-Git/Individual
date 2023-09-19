import pdftotext
import openpyxl
from openpyxl.styles import Font

Inten_Header = ['Analyte', 'Mass', 'Meas. Intens. Mean', 'Meas. Intens. RSD', 'Blank Intensity', 'Blank Intens. RSD']
Concent_Header = ['Analyte', 'Mass', 'Net Intens. Mean', 'Conc. Mean', 'Conc. SD', 'Conc. RSD', 'Sample Unit']
Cali_Header = ['Analyte', 'Mass', 'Curve Type', 'Slope', 'Correlation Coefficient']


def PDF_Read_Start(PDF_Path, Excel_Path, ProgressBar, PDF_Btn, Excel_Btn, Start_Btn):
    global Inten_Header
    global Concent_Header
    global Cali_Header

    PDF_Btn.setEnabled(False)
    Excel_Btn.setEnabled(False)
    Start_Btn.setEnabled(False)
    ProgressBar.setValue(0)

    PDF_File = open(PDF_Path, 'rb')
    FileReader = pdftotext.PDF(PDF_File)
    ProgressBar.setValue(10)

    SampleID_List, newPDF = Find_Same_SampleID_Page_And_Sum(FileReader)
    ProgressBar.setValue(15)

    Summary_Index_List = Get_Pages_Summary_Index(newPDF)
    ProgressBar.setValue(20)

    newPDF_List_After_Summary = Make_Pages_to_List(newPDF, Summary_Index_List)
    ProgressBar.setValue(25)

    Intensity_line_num_list, concentration_line_num_list, calibration_line_num_list = Get_Columns_Row_Index(
        newPDF_List_After_Summary)
    ProgressBar.setValue(30)

    intensity, concentration, calibration = Get_Columns_Text(newPDF_List_After_Summary,
                                                             Intensity_line_num_list,
                                                             concentration_line_num_list,
                                                             calibration_line_num_list)
    ProgressBar.setValue(35)

    inten_head_range, con_head_range, cali_head_range = Get_Header_Index(intensity, concentration, calibration)
    ProgressBar.setValue(40)

    Inten_Columns_Data = Set_Data_to_Proper_Columns_Intensity(intensity, inten_head_range)
    ProgressBar.setValue(45)

    Concent_Columns_Data = Set_Data_to_Proper_Columns_Concent(concentration, con_head_range)
    ProgressBar.setValue(50)

    Cali_Columns_Data = Set_Data_to_Proper_Columns_Cali(calibration, cali_head_range)
    ProgressBar.setValue(65)

    Sorted_Concent_Data = Get_Sorted_Excel_Data_Index(Concent_Columns_Data, 'Conc. Mean')
    ProgressBar.setValue(75)

    # 엑셀 파일 열기
    Result_Excel = openpyxl.load_workbook(Excel_Path)
    ProgressBar.setValue(85)

    # 시트 추가하기
    SheetName_Con = 'Concentration'
    SheetName_Inten = 'Intensities'

    if SheetName_Con not in Result_Excel.sheetnames:
        Result_Excel.create_sheet(SheetName_Con)
    else:
        Result_Excel.remove(Result_Excel[SheetName_Con])
        Result_Excel.create_sheet(SheetName_Con)

    if SheetName_Inten not in Result_Excel.sheetnames:
        Result_Excel.create_sheet(SheetName_Inten)
    else:
        Result_Excel.remove(Result_Excel[SheetName_Inten])
        Result_Excel.create_sheet(SheetName_Inten)


    # 추가된 시트 가져오기
    new_sheet_con = Result_Excel[SheetName_Con]
    new_sheet_inten = Result_Excel[SheetName_Inten]

#==================================================================================================================
    first_column_index = len(Concent_Columns_Data[0]['Analyte'])//2 + 2
    new_sheet_con.cell(row=1, column=first_column_index - 1).value = 'Concentration'  # 컬럼 1에 데이터 삽입
    new_sheet_con.cell(row=1, column=first_column_index + 1).value = 'Conc. Mean'  # 컬럼 1에 데이터 삽입
    # 굵은 글꼴 설정
    bold_font = Font(bold=True)
    con_cell1 = new_sheet_con.cell(row=1, column=first_column_index - 1)
    con_cell2 = new_sheet_con.cell(row=1, column=first_column_index + 1)
    con_cell1.font = bold_font
    con_cell2.font = bold_font

    for i, data in enumerate(Concent_Columns_Data[0]['Analyte']):
        new_sheet_con.cell(row=2, column=(i+3)).value = data  # 컬럼 1에 데이터 삽입

    ProgressBar.setValue(90)

    con_max_length = 0
    for PageIdx in range(len(Concent_Columns_Data)):
        if len(Concent_Columns_Data[PageIdx]) > 0:
            cell_col = PageIdx + 1
            new_sheet_con.cell(row=PageIdx+4, column=1).value = SampleID_List[PageIdx]  # 컬럼 1에 데이터 삽입
            con_max_length = max(con_max_length, len(str(SampleID_List[PageIdx])))

            for i, data in enumerate(Concent_Columns_Data[PageIdx]['Conc. Mean']):
                new_sheet_con.cell(row=PageIdx+4, column=(i+3)).value = data  # 컬럼 1에 데이터 삽입

        else:
            new_sheet_con.cell(row=PageIdx+4, column=1).value = SampleID_List[PageIdx]  # 컬럼 1에 데이터 삽입


#==================================================================================================================

    new_sheet_inten.cell(row=1, column=first_column_index - 1).value = 'Intensities'  # 컬럼 1에 데이터 삽입
    new_sheet_inten.cell(row=1, column=first_column_index + 1).value = 'Meas. Intens. Mean'  # 컬럼 1에 데이터 삽입
    # 굵은 글꼴 설정
    bold_font = Font(bold=True)
    inten_cell1 = new_sheet_inten.cell(row=1, column=first_column_index - 1)
    inten_cell2 = new_sheet_inten.cell(row=1, column=first_column_index + 1)
    inten_cell1.font = bold_font
    inten_cell2.font = bold_font

    for i, data in enumerate(Inten_Columns_Data[0]['Analyte']):
        new_sheet_inten.cell(row=2, column=(i+3)).value = data  # 컬럼 1에 데이터 삽입

    for PageIdx in range(len(Inten_Columns_Data)):
        if len(Inten_Columns_Data[PageIdx]) > 0:
            new_sheet_inten.cell(row=PageIdx+4, column=1).value = SampleID_List[PageIdx]  # 컬럼 1에 데이터 삽입

            for i, data in enumerate(Inten_Columns_Data[PageIdx]['Meas. Intens. Mean']):
                new_sheet_inten.cell(row=PageIdx+4, column=(i+3)).value = data  # 컬럼 1에 데이터 삽입

        else:
            new_sheet_inten.cell(row=PageIdx+4, column=1).value = SampleID_List[PageIdx]  # 컬럼 1에 데이터 삽입

#==================================================================================================================

    ProgressBar.setValue(95)

    # 칼럼 너비 조정
    new_sheet_con.column_dimensions[new_sheet_con.cell(row=1, column=1).column_letter].width = con_max_length
    new_sheet_inten.column_dimensions[new_sheet_inten.cell(row=1, column=1).column_letter].width = con_max_length
    # 1열의 너비를 데이터 길이에 맞게 조정

    # 변경된 내용을 엑셀 파일에 저장
    Result_Excel.save(Excel_Path)
    ProgressBar.setValue(100)

    PDF_Btn.setEnabled(True)
    Excel_Btn.setEnabled(True)
    Start_Btn.setEnabled(True)



def Find_Same_SampleID_Page_And_Sum(PDF):
    New_PDF = []
    SampleID = []
    for PageNum in range(len(PDF) - 1):
        SumBuf = ""
        prev_sample_id = ""
        next_sample_id = ""
        for line in PDF[PageNum].split('\n'):
            if line.lower().startswith('sample id:'):
                prev_sample_id = line.split(':')[1].strip()

        for line in PDF[PageNum + 1].split('\n'):
            if line.lower().startswith('sample id:'):
                next_sample_id = line.split(':')[1].strip()

                if prev_sample_id == next_sample_id:
                    cut_under = PDF[PageNum].split('\n')
                    cut_under = cut_under[:-5]

                    next_page = PDF[PageNum + 1].split('\n')
                    next_page = next_page[:-5]

                    SumBuf = '\n'.join(cut_under + next_page)
                    New_PDF.append(SumBuf)
                    SampleID.append(prev_sample_id)

    return SampleID, New_PDF


def Make_Pages_to_List(PDF, Summary_Index):
    Return_Page = []
    for idx, page in enumerate(PDF):
        lines = page.split('\n')
        lines = lines[Summary_Index[idx]:]
        Return_Page.append(lines)

    return Return_Page

def Get_Pages_Summary_Index(PDF):
    Summary_Start_Index = []

    for newPage in PDF:
        page_lines = newPage.split('\n')

        buf = []

        for idx, text in enumerate(page_lines):
            if 'summary' in text.lower():
                buf.append(idx)

        for summary_num in buf:
            if summary_num != 0:
                Summary_Start_Index.append(summary_num)
                break

    return Summary_Start_Index


def Get_Columns_Row_Index(PDF_Lines_List):
    Find_Columns = ['intensities', 'concentration results', 'calibration']
    intensity_col_num = []
    concentration_col_num = []
    calibration_col_num = []

    for Page_Line in PDF_Lines_List:
        for idx, text in enumerate(Page_Line):
            if Find_Columns[0].lower() in text.lower():
                intensity_col_num.append(idx)
            elif Find_Columns[1].lower() in text.lower():
                concentration_col_num.append(idx)
            elif Find_Columns[2].lower() in text.lower():
                calibration_col_num.append(idx)

    return intensity_col_num, concentration_col_num, calibration_col_num


def Get_Columns_Text(allPages, Intensity_num, Concent_num, Calib_num):
    Intensity_text = []
    Concent_text = []
    Calib_text = []

    for idx, onePage in enumerate(allPages):
        Intensity_text.append(onePage[Intensity_num[idx]: Concent_num[idx]])
        Concent_text.append(onePage[Concent_num[idx]: Calib_num[idx]])

        end_space_count = 0
        for i in range(len(onePage) - 1, -1, -1):
            if onePage[i].strip() == '':
                end_space_count += 1
            else:
                break

        Calib_text.append(onePage[Calib_num[idx]: -(end_space_count - 1)])

    return Intensity_text, Concent_text, Calib_text


def Get_Header_Index(inten, con, cali):
    ALL_Inten_Header_Index = []
    ALL_Concent_Header_Index = []
    ALL_Cali_Header_Index = []

    global Inten_Header
    global Concent_Header
    global Cali_Header

    prev_inten_index = 0
    next_inten_index = 0

    prev_con_index = 0
    next_con_index = 0

    prev_cali_index = 0
    next_cali_index = 0

    for PageIdx in range(len(inten)):
        Inten_Header_Index = []
        for idx in range(len(Inten_Header) - 1):
            prev_inten_index = inten[PageIdx][1].index(Inten_Header[idx])
            next_inten_index = inten[PageIdx][1].index(Inten_Header[idx + 1])
            buf = [prev_inten_index, next_inten_index]
            Inten_Header_Index.append(buf)

        last_inten_buf = [next_inten_index, len(inten[PageIdx])]
        Inten_Header_Index.append(last_inten_buf)
        ALL_Inten_Header_Index.append(Inten_Header_Index)

    for PageIdx in range(len(con)):
        Concent_Header_Index = []
        for idx in range(len(Concent_Header) - 1):
            prev_con_index = con[PageIdx][1].index(Concent_Header[idx])
            next_con_index = con[PageIdx][1].index(Concent_Header[idx + 1])
            buf = [prev_con_index, next_con_index]
            Concent_Header_Index.append(buf)
        last_con_buf = [next_con_index, len(con[PageIdx])]
        Concent_Header_Index.append(last_con_buf)
        ALL_Concent_Header_Index.append(Concent_Header_Index)

    for PageIdx in range(len(cali)):
        Cali_Header_Index = []
        for idx in range(len(Cali_Header) - 1):
            prev_cali_index = cali[PageIdx][1].index(Cali_Header[idx])
            next_cali_index = cali[PageIdx][1].index(Cali_Header[idx + 1])
            buf = [prev_cali_index, next_cali_index]
            Cali_Header_Index.append(buf)

        last_cali_buf = [next_cali_index, len(cali[PageIdx])]
        Cali_Header_Index.append(last_cali_buf)
        ALL_Cali_Header_Index.append(Cali_Header_Index)

    return ALL_Inten_Header_Index, ALL_Concent_Header_Index, ALL_Cali_Header_Index


def Set_Data_to_Proper_Columns_Intensity(inten_data, head_range):
    Intesities_columns_data = []
    global Inten_Header

    for PageIdx in range(len(inten_data)):
        page_data = {
            'Analyte': [],
            'Mass': [],
            'Meas. Intens. Mean': [],
            'Meas. Intens. RSD': [],
            'Blank Intensity': [],
            'Blank Intens. RSD': []
        }

        data = inten_data[PageIdx][2:]
        now_page_head_range = head_range[PageIdx]

        for i in range(len(data)):
            now_lines = data[i].split(' ')
            count = 0

            for data_idx, text in enumerate(now_lines):
                if text != '':
                    count += len(text)

                    for head_idx, number in enumerate(now_page_head_range):
                        start_num = number[0]
                        finish_num = number[1]
                        buf = [x for x in range(start_num, finish_num + 1)]

                        if count in buf:
                            page_data[Inten_Header[head_idx]].append(text)
                            break
                else:
                    count += 1

        Intesities_columns_data.append(page_data)

    return Intesities_columns_data


def Set_Data_to_Proper_Columns_Concent(con_data, head_range):
    Concent_columns_data = []
    global Concent_Header

    for PageIdx in range(len(con_data)):
        page_data = {
            'Analyte': [],
            'Mass': [],
            'Net Intens. Mean': [],
            'Conc. Mean': [],
            'Conc. SD': [],
            'Conc. RSD': [],
            'Sample Unit': []
        }

        data = con_data[PageIdx][2:]
        now_page_head_range = head_range[PageIdx]

        for i in range(len(data)):
            now_lines = data[i].split(' ')
            count = 0

            for data_idx, text in enumerate(now_lines):
                if text != '':
                    count += len(text)

                    for head_idx, number in enumerate(now_page_head_range):
                        start_num = number[0]
                        finish_num = number[1]
                        buf = [x for x in range(start_num, finish_num + 1)]

                        if count in buf:
                            page_data[Concent_Header[head_idx]].append(text)
                            break
                else:
                    count += 1

        Concent_columns_data.append(page_data)

    return Concent_columns_data


def Set_Data_to_Proper_Columns_Cali(cali_data, head_range):
    Cali_columns_data = []
    global Cali_Header

    for PageIdx in range(len(cali_data)):
        page_data = {
            'Analyte': [],
            'Mass': [],
            'Curve Type': [],
            'Slope': [],
            'Correlation Coefficient': []
        }

        data = cali_data[PageIdx][2:]
        now_page_head_range = head_range[PageIdx]

        for i in range(len(data)):
            now_lines = data[i].split(' ')
            count = 0

            for data_idx, text in enumerate(now_lines):
                if text != '':
                    count += len(text)

                    for head_idx, number in enumerate(now_page_head_range):
                        start_num = number[0]
                        finish_num = number[1]
                        buf = [x for x in range(start_num, finish_num + 1)]

                        if count in buf:
                            page_data[Cali_Header[head_idx]].append(text)
                            break
                else:
                    count += 1

        Cali_columns_data.append(page_data)

    return Cali_columns_data


def Get_Sorted_Excel_Data_Index(All_Data, Column):
    All_Return_Data = []
    for PageIdx, PageData in enumerate(All_Data):
        Use_Sorting_Data = PageData[Column]

        if len(Use_Sorting_Data) == 0:
            All_Return_Data.append([])
        else:
            # (숫자, 인덱스) 쌍의 튜플 리스트 생성
            sort_tuples = [(num, idx) for idx, num in enumerate(Use_Sorting_Data)]
            sorted_numbers = sorted(sort_tuples)

            # 정렬된 숫자와 원래 인덱스 출력
            sorted_values = [num for num, _ in sorted_numbers]
            original_indexes = [i for _, i in sorted_numbers]

            Data = {
                'Analyte': [],
                'Conc. Mean': []
            }

            for sort_idx in original_indexes:
                Data['Analyte'].append(All_Data[PageIdx]['Analyte'][sort_idx])
                Data['Conc. Mean'].append(All_Data[PageIdx]['Conc. Mean'][sort_idx])

            All_Return_Data.append(Data)

    return All_Return_Data